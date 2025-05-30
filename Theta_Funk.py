from mpmath import mp, almosteq
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
from scipy.linalg import eigvals









# Имя файла
FILE_NAME = "statistics_Nmax60_k2_error.xlsx"
# Если файла нет - создаем новый, иначе загружаем
if not os.path.exists(FILE_NAME):
    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"
    # Заголовки
    ws.append(["Номер", "двойкаВ", "Знак А", "Знак C", "Знак q", "Период q", "ЗнакSвSigma", "ЗнаквторойчастивSigma", "МнимаяEps", "ЗнакRo", "МнимаяRo", "RoВKдва", "C(X0, 0, Z0)", "C(X1, 0, Z1)", "C0", "C001", "C01", "C1", "ПравильностьC0", "ПравильностьОмбелическойТочки", "АдекватностьXZ"])
    # Делаем заголовки жирными
    for cell in ws[1]:
        cell.font = Font(bold=True)
else:
    wb = load_workbook(FILE_NAME)
    ws = wb.active
    ws.append(["Номер", "двойкаВ", "Знак А", "Знак C", "Знак q", "Период q", "ЗнакSвSigma", "ЗнаквторойчастивSigma", "МнимаяEps", "ЗнакRo", "МнимаяRo", "RoВKдва", "C(X0, 0, Z0)", "C(X1, 0, Z1)", "C0", "C001", "C01", "C1", "ПравильностьC0", "ПравильностьОмбелическойТочки", "АдекватностьXZ"])


# Пример: вычисляем данные и записываем по очереди
def calculate_and_save_metric(Номер, двойкаВ, ЗнакА, ЗнакC, Знакq, ПериодПоq, ЗнакSвSigma, ЗнаквторойчастивSigma, МнимаяEps, ЗнакRo, мнимаяRo, RoВKдва, Cx0z0, Cx1z1, C0, C001, C01, C1, ПравильностьC0, ПравильностьОмбелическойТочки, АдекватностьXZ):
    # Преобразуем комплексные числа в строки или отдельные вещественные/мнимые части
    def process_value(value):
        if hasattr(value, 'imag'):  # Если это комплексное число
            return f"{float(value.real)} + {float(value.imag)}j"
        return value

    # Обрабатываем все значения, которые могут быть комплексными
    processed_data = [
        Номер,
        двойкаВ,
        ЗнакА,
        ЗнакC,
        Знакq,
        ПериодПоq,
        ЗнакSвSigma,
        ЗнаквторойчастивSigma,
        МнимаяEps,
        ЗнакRo,
        мнимаяRo,
        RoВKдва,
        process_value(Cx0z0),
        process_value(Cx1z1),
        process_value(C0),
        process_value(C001),
        process_value(C01),
        process_value(C1),
        ПравильностьC0,
        ПравильностьОмбелическойТочки,
        АдекватностьXZ
    ]

    ws.append(processed_data)
    wb.save(FILE_NAME)


mp.dps = 50  # высокая точность

# Заданные корни (в порядке как на рисунке)
roots = [0, 1, 9]
z1, z2, z3 = roots

c=1
b=2
a=3

g = 2


def P3(z):
    return -z*(z-c**2)*(z-a**2)

def sqrtP(z, branch=1, opt = 0):
    """Корень с выбором ветви (branch=1 для верхнего, -1 для нижнего листа)"""
    return branch * mp.sqrt(P3(z))


# Интеграл по дуге окружности от 0 до pi (верхняя) и от pi до 2pi (нижняя)
def arc_integral(center, radius, func, opt=0, change_brunch = False):
    u = np.linspace(0, 2*np.pi, 101)  # углы
    values = []
    prev_val = None

    for fi in u:
        z = center + radius * mp.exp(1j * fi)
        f1 = func(z)
        f2 = -f1  # альтернативная ветвь

        if prev_val is not None:
            if abs(f1 - prev_val) > abs(f2 - prev_val):
                f1 = f2
        else:
            f1 = func(z)

        if (change_brunch and fi==u[len(u)//2]):
            f1 = -f1

        prev_val = f1
        values.append((z, f1))

        if opt == 1:
            print(f"fi = {float(fi/mp.pi):.2f}Pi,  Z = {z}, f(z) = {f1}")

    # Численное интегрирование по траектории:
    integral = 0
    for i in range(1, len(values)):
        z_prev, f_prev = values[i-1]
        z_curr, f_curr = values[i]
        dz = z_curr - z_prev
        # Среднее значение функции на участке
        avg_f = 0.5 * (f_prev + f_curr)
        integral += avg_f * dz

    return integral

def lin_integral(down, upper, func):
    integral = mp.quad(func, [down, upper])
    return integral


Nmax = 60

def theta_function00(B, z, n_max=Nmax):
    Sum=0
    for M in range(-n_max, n_max):
        Sum += mp.exp(0.5*B*M**2+M*z)
    return Sum


def theta_function01(B, z, n_max=Nmax):
    Sum = mp.exp(B / 8.0 + z / 2.0) * theta_function00(B, z + B / 2, n_max=n_max)
    return Sum


def theta_function10(B, z, n_max=Nmax):
    Sum = theta_function00(B, z + mp.pi * mp.j, n_max=n_max)
    return Sum


def theta_function11(B, z, n_max=Nmax):
    Sum = mp.exp(B/8.0 + z/2.0 + mp.pi*mp.j/2.0 )*theta_function00(B, z + mp.pi*mp.j + B/2.0, n_max=n_max)
    return Sum


def theta_function11D_Summ(B, z, n_max=Nmax):
    Sum = 0
    for M in range(-n_max, n_max):
        Sum += M*mp.exp(0.5 * B * M**2 + M*z + M*mp.pi*mp.j + M*B/2)
    return Sum


def theta_function11D(B, z, n_max=Nmax):
    theta_function11D = 0.5*theta_function11(B, z) + mp.exp(B/8+0.5*mp.pi*mp.j + 0.5*z)*theta_function11D_Summ(B, z, n_max=n_max)
    return theta_function11D


Xomb, Yomb, Zomb = a * mp.sqrt((a ** 2 - b ** 2) / (a ** 2 - c ** 2)), 0, c * mp.sqrt(
    (b ** 2 - c ** 2) / (a ** 2 - c ** 2))


Близость = 1e-2





def check_C(C1, opt="равенство"):
    # Проверяем, что мнимая часть практически нулевая
    if opt == "равенство":
        if almosteq(C1.imag, 0, abs_eps=1e-2):  # Допустимая погрешность (можно настроить)
            real_part = C1.real
            if almosteq(real_part, 1, abs_eps=1e-2):
                return True
        return False
    elif opt == "В интервале":
        if almosteq(C1.imag, 0, abs_eps=1e-2):  # Допустимая погрешность (можно настроить)
            real_part = C1.real
            if 0 <= real_part <= 1:
                return True
        return False



N=0
for двойкаВ in range(0, 2):
    for ЗнакА in range(0,2):
        for ЗнакC in range(0, 2):
            for Знакq in range(0, 2):
                for ПериодПоq in range(0, 2):
                    for ЗнакSвSigma in range(0, 2):
                        for МнимаяSigma in range(0, 2):
                            for ЗнаквторойчастивSigma in range(0, 2):
                                for МнимаяEps in range(0, 2):
                                    for ЗнакRo in range(0, 2):
                                        for мнимаяRo in range(0, 2):
                                            for RoВKдва in range(0, 2):

                                                N+=1





                                                f_int = lambda z: 1.0 / (sqrtP(z, branch=1, opt=1)*mp.pi*mp.j)
                                                Qou = lin_integral(c**2, a**2, f_int)
                                                Aou = lin_integral(0, c**2, f_int)*(-1)**ЗнакА

                                                Bou = (2**двойкаВ)*mp.pi*mp.j*Qou/Aou*(-1)**ЗнакА

                                                if mp.re(Bou)>0: continue

                                                s=np.sqrt(b**2*(b**2-c**2)*(a**2-b**2))

                                                f_C = lambda z: s / (sqrtP(z, branch=1, opt=1)*(z-b**2)*mp.pi*mp.j)
                                                Cou = lin_integral(0, c**2, f_C)*(-1)**ЗнакC

                                                f_q = lambda z: 1.0 / (sqrtP(z, branch=1, opt=1)*Aou)
                                                if ПериодПоq ==0: q=2*lin_integral(c**2, a**2, f_q)*(-1)**Знакq
                                                else: q=2*lin_integral(b**2, a**2, f_q)*(-1)**Знакq
                                                Sigma=s*(-1)**ЗнакSвSigma + Cou*b**2/Aou*(-1)**ЗнаквторойчастивSigma
                                                if МнимаяSigma == 1: Sigma *= mp.j
                                                eps=1
                                                if МнимаяEps==1: eps=mp.j
                                                Ro = -2*b**2/Aou*(-1)**ЗнакRo
                                                if мнимаяRo==1: Ro*=mp.j



                                                k1 = c*mp.sqrt((b**2-c**2)/(a**2-c**2))*theta_function11(Bou, q/2)/theta_function00(Bou, q/2)
                                                #Заменил на b**2
                                                k2 = Aou * b**2 * theta_function11(Bou, q/2)/(theta_function11D(Bou, 0)*(Ro**RoВKдва))
                                                k3 = a*mp.sqrt((a**2-b**2)/(a**2-c**2))*theta_function11(Bou, q/2)/theta_function01(Bou, q/2)




                                                X_tau = lambda tau: k3*(mp.exp(Sigma*tau+eps)*theta_function01(Bou, Ro*tau+q/2.0) - mp.exp(-Sigma*tau-eps)*theta_function01(Bou, Ro*tau-q/2.0)) / (mp.exp(Sigma*tau+eps)*theta_function11(Bou, Ro*tau+q/2) + mp.exp(-Sigma*tau-eps)*theta_function11(Bou, Ro*tau-q/2))

                                                Z_tau = lambda tau: k1*(mp.exp(Sigma*tau+eps)*theta_function00(Bou, Ro*tau+q/2.0) - mp.exp(-Sigma*tau-eps)*theta_function00(Bou, Ro*tau-q/2.0)) / (mp.exp(Sigma*tau+eps)*theta_function11(Bou, Ro*tau+q/2) + mp.exp(-Sigma*tau-eps)*theta_function11(Bou, Ro*tau-q/2))

                                                Y_tau = lambda tau: k2*(theta_function11(Bou, Ro*tau)) / (mp.exp(Sigma*tau+eps)*theta_function11(Bou, Ro*tau+q/2) + mp.exp(-Sigma*tau-eps)*theta_function11(Bou, Ro*tau-q/2))





                                                X0, Y0, Z0 = X_tau(0), Y_tau(0), Z_tau(0)
                                                X001, Y001, Z001 = X_tau(0.01), Y_tau(0.01), Z_tau(0.01)
                                                X01, Y01, Z01 = X_tau(0.1), Y_tau(0.1), Z_tau(0.1)
                                                X1, Y1, Z1 = X_tau(1), Y_tau(1), Z_tau(1)

                                                C = lambda X,Y,Z: X**2/a**2 + Y**2/b**2 + Z**2/c**2

                                                C0, C001, C01, C1 = C(X0, Y0, Z0), C(X001, Y001, Z001), C(X01, Y01, Z01), C(X1, Y1, Z1)

                                                if check_C(C0) and check_C(C001) and check_C(C01) and check_C(C1):
                                                    ПравильностьC0 = "Правильно!"
                                                    print("\n\n\n\nПравильно!\n\n\n\n")
                                                else: ПравильностьC0 = "неправильно"

                                                if abs(X0-Xomb) < 1e-3 and abs(Y0-Yomb) < 1e-3 and abs(Z0-Zomb) < 1e-3: ПравильностьОмбелическойТочки = "Правильная омбилическая точка!"
                                                else: ПравильностьОмбелическойТочки = "неправильная омбилическая точка"

                                                if check_C(C(X1, 0, Z1), "В интервале"):
                                                    АдекватностьXZ = "X1Z1 адекватные"
                                                    if ПравильностьОмбелическойТочки == "Правильная омбилическая точка!":
                                                        from mpmath import mp

                                                        # Предполагаем, что все переменные — mpf/mpc, кроме, возможно, N
                                                        print(
                                                            f"N = {N}, A = {mp.nstr(Aou, 5)}, B = {mp.nstr(Bou, 5)}, C = {mp.nstr(Cou, 5)}, "
                                                            f"s = {mp.nstr(s, 5)}, q = {mp.nstr(q, 5)}, Sigma = {mp.nstr(Sigma, 5)}, "
                                                            f"Ro = {mp.nstr(Ro, 5)}, k1 = {mp.nstr(k1, 5)}, k2 = {mp.nstr(k2, 5)}, "
                                                            f"k3 = {mp.nstr(k3, 5)}, X(1) = {mp.nstr(X1, 5)}, Y(1) = {mp.nstr(Y1, 5)}, "
                                                            f"Z(1) = {mp.nstr(Z1, 5)}")
                                                else: АдекватностьXZ = "X1Z1 не соответствуют эллипсоиду"

                                                calculate_and_save_metric(N, двойкаВ, ЗнакА, ЗнакC,
                                                                          Знакq, ПериодПоq, ЗнакSвSigma, ЗнаквторойчастивSigma,
                                                                          МнимаяEps, ЗнакRo, мнимаяRo, RoВKдва, C(X0, 0, Z0), C(X1, 0, Z1), C0, C001, C01,
                                                                          C1, ПравильностьC0, ПравильностьОмбелическойТочки, АдекватностьXZ)



    wb.close()







